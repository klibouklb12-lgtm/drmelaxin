#!/usr/bin/env python3
"""
Dr.Melaxin — Pre-formatted Google Sheet generator
Creates a beautiful .xlsx file ready to upload to Google Sheets.

4 Tabs:
1. Dashboard  — Arabic statistics (KPIs, status breakdown, top wilayas, recent orders)
2. Orders     — English headers, colored status dropdown
3. Stock      — Simple, elegant
4. Product    — Product settings

Output: /home/z/my-project/download/DrMelaxin-Sheet.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ============================================================================
#  COLORS (match the Apps Script v3)
# ============================================================================
COLOR_BURGUNDY = "8B1538"
COLOR_GOLD = "D4AF37"
COLOR_CREAM = "F7F5F2"
COLOR_DARK = "2A2520"
COLOR_MUTED = "6B6358"
COLOR_WHITE = "FFFFFF"

STATUS_COLORS = {
    "New": "3080FF",
    "Confirmed": "016630",
    "Shipped": "FFD700",
    "Delivered": "2F7D5B",
    "Cancelled": "E40014",
}
STATUS_TEXT_COLORS = {
    "New": "FFFFFF",
    "Confirmed": "FFFFFF",
    "Shipped": "1C1815",
    "Delivered": "FFFFFF",
    "Cancelled": "FFFFFF",
}

# ============================================================================
#  HELPERS
# ============================================================================
def thin_border():
    side = Side(style="thin", color="D0D0D0")
    return Border(left=side, right=side, top=side, bottom=side)

def header_style(cell):
    cell.font = Font(name="Arial", bold=True, color=COLOR_WHITE, size=11)
    cell.fill = PatternFill(start_color=COLOR_DARK, end_color=COLOR_DARK, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

def title_style(cell):
    cell.font = Font(name="Arial", bold=True, color=COLOR_WHITE, size=24)
    cell.fill = PatternFill(start_color=COLOR_BURGUNDY, end_color=COLOR_BURGUNDY, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# ============================================================================
#  CREATE WORKBOOK
# ============================================================================
wb = Workbook()
wb.properties.creator = "Dr.Melaxin"

# Remove default sheet (we'll create our own)
default = wb.active
wb.remove(default)

# ============================================================================
#  1. DASHBOARD TAB
# ============================================================================
d = wb.create_sheet("Dashboard")
d.sheet_view.showGridLines = False

# Column widths
d.column_dimensions["A"].width = 3
d.column_dimensions["B"].width = 22
d.column_dimensions["C"].width = 22
d.column_dimensions["D"].width = 22
d.column_dimensions["E"].width = 22
d.column_dimensions["F"].width = 3

# Row 2: Title
d.merge_cells("B2:E2")
d["B2"] = "لوحة التحكم · Dashboard"
title_style(d["B2"])
d.row_dimensions[2].height = 50

# Row 4: Section header
d["B4"] = "📊 النظرة العامة · Overview"
d["B4"].font = Font(name="Arial", bold=True, color=COLOR_BURGUNDY, size=14)

# Rows 5-7: KPI Cards
kpis = [
    ("B", "إجمالي الطلبات", "Total Orders", "=COUNTA(Orders!B2:B)", COLOR_BURGUNDY),
    ("C", "إجمالي الإيرادات", "Revenue (DA)", "=SUM(Orders!I2:I)", COLOR_GOLD),
    ("D", "متوسط الطلب", "Avg Order (DA)", "=IFERROR(ROUND(AVERAGE(Orders!I2:I),0),0)", "016630"),
    ("E", "المخزون الحالي", "Current Stock", "=Stock!B2", "2F7D5B"),
]

for col, ar_label, en_label, formula, color in kpis:
    # Row 5: Arabic label (colored background)
    cell = d[f"{col}5"]
    cell.value = ar_label
    cell.font = Font(name="Arial", bold=True, color=COLOR_WHITE, size=11)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

    # Row 6: Big number (formula)
    cell = d[f"{col}6"]
    cell.value = formula
    cell.font = Font(name="Arial", bold=True, color=color, size=28)
    cell.fill = PatternFill(start_color=COLOR_CREAM, end_color=COLOR_CREAM, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "#,##0"
    cell.border = thin_border()

    # Row 7: English label
    cell = d[f"{col}7"]
    cell.value = en_label
    cell.font = Font(name="Arial", color=COLOR_MUTED, size=9)
    cell.fill = PatternFill(start_color=COLOR_CREAM, end_color=COLOR_CREAM, fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border()

d.row_dimensions[5].height = 24
d.row_dimensions[6].height = 50
d.row_dimensions[7].height = 18

# Row 9: Status breakdown header
d["B9"] = "📋 حالة الطلبات · Order Status"
d["B9"].font = Font(name="Arial", bold=True, color=COLOR_BURGUNDY, size=14)

# Row 10: Headers
d["B10"] = "Status"
d["C10"] = "Count"
d["D10"] = "العدد"
for col in ["B", "C", "D"]:
    header_style(d[f"{col}10"])
d.row_dimensions[10].height = 28

# Rows 11-15: Status rows
status_arabic = {
    "New": "جديد",
    "Confirmed": "مؤكد",
    "Shipped": "مشحون",
    "Delivered": "مسلّم",
    "Cancelled": "ملغى",
}

for i, status in enumerate(["New", "Confirmed", "Shipped", "Delivered", "Cancelled"]):
    row = 11 + i
    # Status (colored)
    cell = d[f"B{row}"]
    cell.value = status
    cell.font = Font(name="Arial", bold=True, color=STATUS_TEXT_COLORS[status], size=11)
    cell.fill = PatternFill(start_color=STATUS_COLORS[status], end_color=STATUS_COLORS[status], fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

    # Count (formula)
    cell = d[f"C{row}"]
    cell.value = f'=COUNTIF(Orders!C2:C,"{status}")'
    cell.font = Font(name="Arial", bold=True, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

    # Arabic name
    cell = d[f"D{row}"]
    cell.value = status_arabic[status]
    cell.font = Font(name="Arial", color=COLOR_MUTED, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

    d.row_dimensions[row].height = 28

# Row 17: Top wilayas header
d["B17"] = "🌍 أعلى الولايات · Top Wilayas"
d["B17"].font = Font(name="Arial", bold=True, color=COLOR_BURGUNDY, size=14)

# Row 18: Headers
d["B18"] = "Wilaya"
d["C18"] = "Orders"
d["D18"] = "الطلبات"
for col in ["B", "C", "D"]:
    header_style(d[f"{col}18"])
d.row_dimensions[18].height = 28

# Rows 19-23: Top 5 wilayas — single QUERY outputs 2 columns (name + count)
d["B19"] = '=IFERROR(QUERY(Orders!F2:I,"select F, count(F) where F is not null and F != \'\' group by F order by count(F) desc limit 5 label F \'Wilaya\', count(F) \'Orders\'",0),"No orders yet")'

for i in range(5):
    row = 19 + i
    for col in ["B", "C", "D"]:
        cell = d[f"{col}{row}"]
        cell.font = Font(name="Arial", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        if col == "C":
            cell.font = Font(name="Arial", bold=True, size=11)
    d.row_dimensions[row].height = 24

# Row 25: Recent orders header
d["B25"] = "🕒 أحدث الطلبات · Recent Orders"
d["B25"].font = Font(name="Arial", bold=True, color=COLOR_BURGUNDY, size=14)

# Row 26: Headers
recent_headers = ["Order No", "Customer", "Wilaya", "Total (DA)"]
for i, h in enumerate(recent_headers):
    cell = d.cell(row=26, column=2 + i, value=h)
    header_style(cell)
d.row_dimensions[26].height = 28

# Rows 27-31: Last 5 orders — single QUERY outputs 4 columns
d.cell(row=27, column=2).value = '=IFERROR(QUERY(Orders!B2:I,"select B, D, F, I where B is not null order by A desc limit 5 label B \'\', D \'\', F \'\', I \'\'",0),"No orders yet")'

for i in range(5):
    row = 27 + i
    for col in range(2, 6):
        cell = d.cell(row=row, column=col)
        cell.font = Font(name="Arial", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        if col == 5:
            cell.font = Font(name="Arial", bold=True, size=11)
    d.row_dimensions[row].height = 24

# Row 33-34: Footer
d.merge_cells("B33:E33")
d["B33"] = "💡 يتم تحديث الإحصائيات تلقائياً عند كل طلب جديد"
d["B33"].font = Font(name="Arial", color=COLOR_MUTED, size=10)
d["B33"].alignment = Alignment(horizontal="center")

d.merge_cells("B34:E34")
d["B34"] = "💡 Statistics auto-update on every new order"
d["B34"].font = Font(name="Arial", color=COLOR_MUTED, size=10)
d["B34"].alignment = Alignment(horizontal="center")

# Background for spacer columns
for row in range(1, 35):
    d.cell(row=row, column=1).fill = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")
    d.cell(row=row, column=6).fill = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")

# ============================================================================
#  2. ORDERS TAB
# ============================================================================
o = wb.create_sheet("Orders")

# Headers (English)
headers = ["Date", "Order No", "Status", "Customer", "Phone", "Wilaya", "Commune", "Qty", "Total (DA)", "Notes", "Idempotency Key"]
for i, h in enumerate(headers):
    cell = o.cell(row=1, column=i + 1, value=h)
    header_style(cell)
o.row_dimensions[1].height = 28

# Column widths
widths = [20, 18, 14, 22, 16, 16, 16, 8, 14, 30, 22]
for i, w in enumerate(widths):
    o.column_dimensions[get_column_letter(i + 1)].width = w

# Freeze header
o.freeze_panes = "A2"

# Auto-filter
o.auto_filter.ref = f"A1:K1"

# Data validation: Status dropdown
status_dv = DataValidation(
    type="list",
    formula1='"New,Confirmed,Shipped,Delivered,Cancelled"',
    allow_blank=False,
    showDropDown=False,  # False = show dropdown arrow
)
status_dv.add(f"C2:C5000")
o.add_data_validation(status_dv)

# Conditional formatting: Status colors
for status, bg_color in STATUS_COLORS.items():
    text_color = STATUS_TEXT_COLORS[status]
    rule = FormulaRule(
        formula=[f'$C2="{status}"'],
        fill=PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid"),
        font=Font(name="Arial", bold=True, color=text_color, size=11),
    )
    o.conditional_formatting.add(f"C2:C5000", rule)

# Default "New" in first 50 rows (light gray)
for row in range(2, 52):
    cell = o.cell(row=row, column=3, value="New")
    cell.font = Font(name="Arial", color="999999", size=11)
    cell.alignment = Alignment(horizontal="center")

# ============================================================================
#  3. STOCK TAB
# ============================================================================
stk = wb.create_sheet("Stock")

# Headers
stk["A1"] = "Product"
stk["B1"] = "Stock"
header_style(stk["A1"])
header_style(stk["B1"])
stk.row_dimensions[1].height = 28

# Product row
stk["A2"] = "Cemenrete CX"
stk["B2"] = 100

# Style product row
for col in ["A", "B"]:
    cell = stk[f"{col}2"]
    cell.font = Font(name="Arial", bold=True, color=COLOR_BURGUNDY, size=14)
    cell.fill = PatternFill(start_color=COLOR_CREAM, end_color=COLOR_CREAM, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

stk.column_dimensions["A"].width = 25
stk.column_dimensions["B"].width = 15

# Instructions (column D)
stk["D1"] = "📦 Stock Management · إدارة المخزون"
stk["D1"].font = Font(name="Arial", bold=True, color=COLOR_BURGUNDY, size=14)

instructions = [
    "",
    "1. Set stock number in B2 (e.g. 100)",
    "2. When you mark order 'Confirmed' → stock auto-reduces",
    "3. Stock ≤ 3 → website shows 'low stock' warning",
    "4. Stock = 0 → website disables ordering",
    "5. To restock → just change B2",
    "",
    "💡 Stock auto-updates via trigger (onEditTrigger)",
    "   when you change Status to 'Confirmed' in Orders tab",
]

for i, text in enumerate(instructions):
    cell = stk.cell(row=2 + i, column=4, value=text)
    if text.startswith("💡"):
        cell.font = Font(name="Arial", color=COLOR_GOLD, bold=True, size=11)
    elif text:
        cell.font = Font(name="Arial", color=COLOR_MUTED, size=11)

stk.column_dimensions["D"].width = 50

# ============================================================================
#  4. PRODUCT TAB
# ============================================================================
p = wb.create_sheet("Product")

# Headers
p["A1"] = "Setting"
p["B1"] = "Value"
header_style(p["A1"])
header_style(p["B1"])
p.row_dimensions[1].height = 28

# Product defaults
product_data = [
    ("brandName", "Dr.Melaxin"),
    ("lineName", "Cemenrete CX"),
    ("subtitle", "Calcium Volume Multi Balm"),
    ("basePrice", 3900),
    ("oldPrice", 5800),
    ("taglineArabic", "✨ ستيك للعناية بالتجاعيد لبشرة أكثر نعومة وتماسكاً."),
    ("descriptionArabic", "تساعد تركيبته على تقليل مظهر التجاعيد والخطوط الدقيقة، مع ترطيب البشرة ومنحها مظهراً أكثر إشراقاً ونعومة."),
    ("benefitsArabic", "مضاد للتجاعيد • تماسك • ترطيب • إشراقة"),
    ("taglineFrench", "✨ Le stick anti-ridules pour une peau visiblement plus lisse et plus ferme."),
    ("descriptionFrench", "Sa formule aide à réduire l'apparence des rides et ridules, tout en apportant hydratation, confort et éclat à la peau."),
    ("benefitsFrench", "Anti-ridules • Fermeté • Hydratation • Éclat"),
    ("badgeArabic", "مضاد للتجاعيد"),
    ("freeShipping", True),
]

for i, (key, value) in enumerate(product_data):
    row = 2 + i
    p.cell(row=row, column=1, value=key)
    p.cell(row=row, column=2, value=value)

    # Style: key column
    p.cell(row=row, column=1).font = Font(name="Arial", bold=True, color=COLOR_BURGUNDY, size=11)
    p.cell(row=row, column=1).fill = PatternFill(start_color=COLOR_CREAM, end_color=COLOR_CREAM, fill_type="solid")
    p.cell(row=row, column=1).border = thin_border()

    # Style: value column
    p.cell(row=row, column=2).font = Font(name="Arial", size=11)
    p.cell(row=row, column=2).fill = PatternFill(start_color=COLOR_CREAM, end_color=COLOR_CREAM, fill_type="solid")
    p.cell(row=row, column=2).border = thin_border()

p.column_dimensions["A"].width = 25
p.column_dimensions["B"].width = 50
p.freeze_panes = "A2"

# ============================================================================
#  SAVE
# ============================================================================
output_path = "/home/z/my-project/download/DrMelaxin-Sheet.xlsx"
wb.save(output_path)

print(f"✅ Created: {output_path}")

# Verify
import os
size = os.path.getsize(output_path)
print(f"   Size: {size:,} bytes ({size/1024:.1f} KB)")
print(f"   Tabs: {wb.sheetnames}")
